import pandas as pd
import openpyxl
from pathlib import Path
# 從 openpyxl 匯入 Workbook, Worksheet 型別，專門用來做型別提示
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

def generate_sales_report(csv_path: Path, template_path: Path, output_path: Path) -> None:
    """
    讀取銷售資料 CSV，依業務單位加總後，填入指定的 Excel 報告範本欄位中。

    """
    # ---------------------------------------------------------
    # 1. 讀取與計算資料 (使用 pandas)
    # ---------------------------------------------------------
    # 讀取 CSV 檔案, 轉為df物件
    df = pd.read_csv(csv_path)
    
    # 依「業務單位」分組，並計算「銷售金額」與「銷售數量」的總和
    grouped = df.groupby('業務單位')[['銷售金額', '銷售數量']].sum().reset_index()
    
    # 【優雅技巧】將 DataFrame 轉換為巢狀字典格式，方便後續以 O(1) 的複雜度快速查閱
    # 轉換後格式範例：{'業務1': {'銷售金額': 12345, '銷售數量': 67}, ...}
    data_dict = grouped.set_index('業務單位').to_dict('index')

    # ---------------------------------------------------------
    # 2. 載入並填寫 Excel 範本 (使用 openpyxl)
    # ---------------------------------------------------------
    # 載入含有特定格式與合併儲存格的 Excel 範本
    wb:Workbook = openpyxl.load_workbook(template_path)
    ws:Worksheet = wb.active

    # 遍歷 Excel 中的每一列 (Row) 來尋找對應的業務單位
    for row in range(1, ws.max_row + 1):
        # 讀取 B 欄 (column=2) 的值
        cell_val = ws.cell(row=row, column=2).value
        
        # 若該儲存格的值剛好是我們的「業務單位」(例如：業務1、業務2)
        if cell_val in data_dict:
            # 將「銷售金額」寫入 C 欄 (column=3)
            ws.cell(row=row, column=3).value = data_dict[cell_val]['銷售金額']
            # 將「銷售數量」寫入 D 欄 (column=4)
            ws.cell(row=row, column=4).value = data_dict[cell_val]['銷售數量']

    # ---------------------------------------------------------
    # 3. 儲存結果
    # ---------------------------------------------------------
    wb.save(output_path)
    wb.close()
    print(f"✅ 報告已成功生成並儲存至：{output_path}")

# ==========================================
# 執行區域
# ==========================================
if __name__ == "__main__":
    # 【優雅技巧】使用 pathlib 處理路徑，取代傳統的 os.path 字串拼接
    BASE_DIR = Path.cwd()
    work_dir = BASE_DIR / '02_data_analysis' / 'outputs'
    print(work_dir)
    
    # 運用 / 運算子優雅地串接路徑
    csv_file = work_dir / 'sales.csv'
    template_file = work_dir / '業務分析報告表格.xlsx'
    output_file = work_dir / '業務分析報告結果.xlsx'

    # 執行函式
    generate_sales_report(csv_file, template_file, output_file)