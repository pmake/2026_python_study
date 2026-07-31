import sys
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 建立基礎路徑 (相容 PyInstaller 打包與一般 Python 腳本執行)
if getattr(sys, 'frozen', False):
    # PyInstaller 打包後的執行檔所在目錄
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    # 一般 Python 腳本執行目錄
    BASE_DIR = Path(__file__).resolve().parent

OUTPUT_DIR = BASE_DIR / 'files'

# 自動建立所需資料夾結構
DIR_SETTING = OUTPUT_DIR / '系統設定'
DIR_SERVICE = OUTPUT_DIR / '系統撈出的當月工時放這'
DIR_TEMPLATE = OUTPUT_DIR / '薪資單表單'
DIR_PAYROLL = OUTPUT_DIR / '產出統計報表'

DIR_SETTING.mkdir(parents=True, exist_ok=True)
DIR_SERVICE.mkdir(parents=True, exist_ok=True)
DIR_TEMPLATE.mkdir(parents=True, exist_ok=True)
DIR_PAYROLL.mkdir(parents=True, exist_ok=True)

INPUT_SETTING = DIR_SETTING / '好寶寶系統設定.xlsx'
INPUT_SERVICE = DIR_SERVICE / '服務紀錄總表.xlsx'
INPUT_TEMPLATE = DIR_TEMPLATE / '好寶寶薪資格式.xlsx'
OUTPUT_PAYROLL = DIR_PAYROLL / '薪資單工時統計表.xlsx'


def check_input_files_exist():
    """檢查必要的輸入 Excel 檔案是否存在"""
    missing = []
    if not INPUT_SETTING.exists():
        missing.append(f" - {INPUT_SETTING}")
    if not INPUT_SERVICE.exists():
        missing.append(f" - {INPUT_SERVICE}")
    if not INPUT_TEMPLATE.exists():
        missing.append(f" - {INPUT_TEMPLATE}")
    
    if missing:
        print("\n【警告】找不到必要的輸入 Excel 檔案：")
        for m in missing:
            print(m)
        print("\n請將對應的 Excel 檔案放入上述目錄後重新執行。")
        return False
    return True


def load_input_data():
    """讀取員工總表、假日定義與服務紀錄總表"""
    # 讀取員工總表
    df_emp = pd.read_excel(INPUT_SETTING, sheet_name='員工總表')
    # 讀取假日定義
    df_hol = pd.read_excel(INPUT_SETTING, sheet_name='假日定義')
    # 讀取服務紀錄總表
    df_srv = pd.read_excel(INPUT_SERVICE, sheet_name='Sheet1')
    
    return df_emp, df_hol, df_srv


def clean_and_process_data(df_emp, df_hol, df_srv):
    """
    清洗資料、轉場加時獨立計算、每日工時分段（以分鐘為單位），並收集異常紀錄。
    """
    anomalies = []
    
    # 建立員工資訊字典與身分證對照
    emp_dict = {}
    for _, row in df_emp.iterrows():
        name = str(row['員工']).strip()
        emp_id = str(row['身分證字號']).strip()
        emp_dict[name] = {
            '姓名': name,
            '員工編號': str(row['員工編號']).strip(),
            '出生日期': str(row['出生日期']).strip(),
            '身分證字號': emp_id,
            '職稱': str(row['職稱']).strip(),
            '工作類別': str(row['工作類別']).strip(),
            'sheet_name': f"{name}_{emp_id}"
        }
        
    # 建立假日 Date Set (YYYY-MM-DD)
    holiday_set = set()
    for _, row in df_hol.iterrows():
        dt_val = pd.to_datetime(row['日期']).strftime('%Y-%m-%d')
        holiday_set.add(dt_val)
        
    # 處理服務紀錄
    df_srv['排班日期_str'] = pd.to_datetime(df_srv['排班日期']).dt.strftime('%Y-%m-%d')
    
    # 數值轉換服務時間長度 (分)
    numeric_duration = pd.to_numeric(df_srv['服務時間長度(分)'], errors='coerce')
    
    # 找出異常格式 (如 "暫無資料" 或 NaN)
    invalid_mask = numeric_duration.isna() | (numeric_duration <= 0)
    invalid_rows = df_srv[invalid_mask]
    
    for _, row in invalid_rows.iterrows():
        nanny_name = str(row['保母姓名']).strip()
        emp_info = emp_dict.get(nanny_name, {})
        emp_id = emp_info.get('身分證字號', '')
        date_str = str(row['排班日期_str'])
        raw_val = row['服務時間長度(分)']
        
        anomalies.append({
            '異常原因': f'資料格式異常(非數值工時: "{raw_val}")',
            '員工姓名': nanny_name,
            '身份證': emp_id,
            '異常日期': date_str
        })
        
    # 留下合法服務紀錄
    valid_srv = df_srv[~invalid_mask].copy()
    valid_srv['duration_min'] = numeric_duration[~invalid_mask].astype(int)
    
    # 統計每位員工的每日數據
    # 結構: daily_records[nanny_name][date_str] = {...}
    daily_records = {name: {} for name in emp_dict}
    
    # 群組化依 保母姓名 與 排班日期
    grouped = valid_srv.groupby(['保母姓名', '排班日期_str'])
    
    for (nanny_name, date_str), group in grouped:
        if nanny_name not in emp_dict:
            continue
            
        emp_info = emp_dict[nanny_name]
        emp_type = emp_info['工作類別']
        is_holiday = date_str in holiday_set
        
        # 1. 轉場加時計算 (按排班時間排序)
        sorted_group = group.sort_values(by='排班時間')
        addresses = sorted_group['居住地址'].tolist()
        
        transit_count = 0
        for i in range(1, len(addresses)):
            if addresses[i] != addresses[i - 1]:
                transit_count += 1
                
        transit_minutes = transit_count * 15
        service_minutes = sorted_group['duration_min'].sum()
        
        # 轉場工時完全獨立統計，不併入當日總工時加總計算 (僅採計照顧服務時間)
        total_daily_minutes = service_minutes
        
        # 2. 轉場加時欄位歸類 (平日轉場加時 vs 假日轉場加時)
        weekday_transit = 0 if is_holiday else transit_minutes
        weekend_transit = transit_minutes if is_holiday else 0
        
        # 3. 當日服務工時分段切片 (單位: 分鐘，上限 12 小時 = 720 分鐘)
        billable_minutes = min(total_daily_minutes, 720)
        
        norm_min = 0
        ot_9_10_min = 0
        ot_11_12_min = 0
        hol_0_2_min = 0
        hol_3_8_min = 0
        hol_9_12_min = 0
        
        if emp_type == '假日班':
            if is_holiday:
                hol_0_2_min = min(billable_minutes, 120)
                if billable_minutes > 120:
                    hol_3_8_min = min(billable_minutes - 120, 360)
                if billable_minutes > 480:
                    hol_9_12_min = min(billable_minutes - 480, 240)
            else:
                # 假日班於平日出勤 (異常排班紀錄)
                norm_min = billable_minutes
                anomalies.append({
                    '異常原因': '假日班於平日出勤',
                    '員工姓名': nanny_name,
                    '身份證': emp_info['身分證字號'],
                    '異常日期': date_str
                })
        else:  # 常日班
            if is_holiday:
                hol_0_2_min = min(billable_minutes, 120)
                if billable_minutes > 120:
                    hol_3_8_min = min(billable_minutes - 120, 360)
                if billable_minutes > 480:
                    hol_9_12_min = min(billable_minutes - 480, 240)
            else:
                norm_min = min(billable_minutes, 480)
                if billable_minutes > 480:
                    ot_9_10_min = min(billable_minutes - 480, 120)
                if billable_minutes > 600:
                    ot_11_12_min = min(billable_minutes - 600, 120)
                    
        daily_records[nanny_name][date_str] = {
            '正常工時': norm_min,
            '平日加班_9_10h': ot_9_10_min,
            '平日加班_11_12h': ot_11_12_min,
            '平日轉場加時': weekday_transit,
            '假日工時_0_2h': hol_0_2_min,
            '假日工時_3_8h': hol_3_8_min,
            '假日工時_9_12h': hol_9_12_min,
            '假日轉場加時': weekend_transit,
        }
        
    # 計算來源排班紀錄之動態連續日曆時間軸 all_dates
    if not df_srv.empty:
        srv_dates = pd.to_datetime(df_srv['排班日期']).dropna()
        min_date = srv_dates.min()
        max_date = srv_dates.max()
        all_dates = [d.strftime('%Y-%m-%d') for d in pd.date_range(min_date, max_date)]
    else:
        all_dates = []

    raw_total_service_hours = valid_srv['duration_min'].sum() / 60.0
    return emp_dict, daily_records, anomalies, raw_total_service_hours, all_dates


def build_payroll_excel(emp_dict, daily_records, anomalies, raw_total_service_hours, all_dates):
    """使用 openpyxl 建立薪資單工時統計表.xlsx 包含四種工作表"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # 樣式設定
    font_header = Font(name='微軟正黑體', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    font_data = Font(name='微軟正黑體', size=10)
    font_check_ok = Font(name='微軟正黑體', size=10, color='006100', bold=True)
    fill_check_ok = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 1. 產出個別員工明細 Sheet (全動態連續日曆數據)
    detail_headers = [
        '日期', '員工類別(常日班/假日班)', '正常工時時數(0~8小時)',
        '平日加班時數(9~10小時)', '平日加班時數(11~12小時)', '平日轉場加時',
        '假日工時(0~2小時)', '假日工時(3~8小時)', '假日工時(9~12小時)', '假日轉場加時'
    ]
    
    stat_last_day_row = 1 + len(all_dates)
    
    for nanny_name, emp_info in emp_dict.items():
        sheet_title = emp_info['sheet_name']
        ws = wb.create_sheet(title=sheet_title)
        
        # 標頭
        ws.append(detail_headers)
        for col_num in range(1, len(detail_headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            
        emp_records = daily_records.get(nanny_name, {})
        emp_type = emp_info['工作類別']
        
        # 補全連續日曆天數
        for date_str in all_dates:
            day_data = emp_records.get(date_str, {
                '正常工時': 0, '平日加班_9_10h': 0, '平日加班_11_12h': 0,
                '平日轉場加時': 0, '假日工時_0_2h': 0, '假日工時_3_8h': 0, '假日工時_9_12h': 0, '假日轉場加時': 0
            })
            
            row_val = [
                date_str,
                emp_type,
                day_data['正常工時'],
                day_data['平日加班_9_10h'],
                day_data['平日加班_11_12h'],
                day_data['平日轉場加時'],
                day_data['假日工時_0_2h'],
                day_data['假日工時_3_8h'],
                day_data['假日工時_9_12h'],
                day_data['假日轉場加時']
            ]
            ws.append(row_val)
            
        # 格式與邊界調整
        for row_idx in range(2, stat_last_day_row + 1):
            ws.cell(row=row_idx, column=1).alignment = align_center
            ws.cell(row=row_idx, column=2).alignment = align_center
            for c_idx in range(1, 11):
                cell = ws.cell(row=row_idx, column=c_idx)
                cell.font = font_data
                cell.border = thin_border
                if c_idx >= 3:
                    cell.number_format = '#,##0'
                    cell.alignment = align_right

    # 2. 產出 員工工時統計表 Sheet
    ws_stat = wb.create_sheet(title='員工工時統計表')
    stat_headers = [
        '員工姓名', '職稱', '身份證字號', '出生年月日', '員工類別(常日班/假日班)',
        '正常工時時數(0~8小時)', '平日加班時數(9~10小時)', '平日加班時數(11~12小時)',
        '平日轉場加時', '假日工時(0~2小時)', '假日工時(3~8小時)', '假日工時(9~12小時)', '假日轉場加時'
    ]
    ws_stat.append(stat_headers)
    for col_num in range(1, len(stat_headers) + 1):
        cell = ws_stat.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        
    stat_row_map = {}
    row_counter = 2
    for nanny_name, emp_info in emp_dict.items():
        s_name = emp_info['sheet_name']
        stat_row_map[nanny_name] = row_counter
        
        # 使用 /60 Excel 公式轉換為小時
        row_val = [
            emp_info['姓名'],
            emp_info['職稱'],
            emp_info['身分證字號'],
            emp_info['出生日期'],
            emp_info['工作類別'],
            f"=SUM('{s_name}'!C2:C{stat_last_day_row})/60",  # 正常工時
            f"=SUM('{s_name}'!D2:D{stat_last_day_row})/60",  # 平日加班 9-10h
            f"=SUM('{s_name}'!E2:E{stat_last_day_row})/60",  # 平日加班 11-12h
            f"=SUM('{s_name}'!F2:F{stat_last_day_row})/60",  # 平日轉場加時
            f"=SUM('{s_name}'!G2:G{stat_last_day_row})/60",  # 假日工時 0-2h
            f"=SUM('{s_name}'!H2:H{stat_last_day_row})/60",  # 假日工時 3-8h
            f"=SUM('{s_name}'!I2:I{stat_last_day_row})/60",  # 假日工時 9-12h
            f"=SUM('{s_name}'!J2:J{stat_last_day_row})/60"   # 假日轉場加時
        ]
        ws_stat.append(row_val)
        
        for c_idx in range(1, 14):
            cell = ws_stat.cell(row=row_counter, column=c_idx)
            cell.font = font_data
            cell.border = thin_border
            if c_idx <= 5:
                cell.alignment = align_center
            else:
                cell.number_format = '0.00'
                cell.alignment = align_right
        row_counter += 1

    # 3. 產出 驗證用資料表 Sheet
    ws_val = wb.create_sheet(title='驗證用資料表')
    val_headers = ['驗證項目', '數值 (小時)', '計算說明與來源', '驗證狀態']
    ws_val.append(val_headers)
    for col_num in range(1, len(val_headers) + 1):
        cell = ws_val.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        
    last_stat_row = 1 + len(emp_dict)
    val_rows = [
        ['1. 原始服務紀錄總工時', raw_total_service_hours, '來自《服務紀錄總表.xlsx》合法服務工時總計', ''],
        ['2. 統計表服務工時總計', f"=SUM('員工工時統計表'!F2:H{last_stat_row})+SUM('員工工時統計表'!J2:L{last_stat_row})", '統計表服務工時欄位 (F:H, J:L) 之全體員工總計', ''],
        ['3. 統計表轉場工時總計', f"=SUM('員工工時統計表'!I2:I{last_stat_row})+SUM('員工工時統計表'!M2:M{last_stat_row})", '統計表平日轉場加時(I欄) + 假日轉場加時(M欄)', ''],
        ['4. 統計表總出勤工時 (服務+轉場)', "=B3+B4", '統計表服務工時 加總 轉場工時 (B3 + B4)', ''],
        ['5. 比對結果與差值 (原始 vs 統計服務工時)', "=B2-B3", '原始服務總工時與統計表服務工時之差值 (B2 - B3)', '=IF(ROUND(ABS(B6),4)=0, "相符 (OK)", "不符 (請檢查)")']
    ]
    
    for r_idx, row_data in enumerate(val_rows, start=2):
        ws_val.append(row_data)
        for c_idx in range(1, 5):
            cell = ws_val.cell(row=r_idx, column=c_idx)
            cell.border = thin_border
            if c_idx == 1 or c_idx == 3:
                cell.font = font_data
                cell.alignment = align_left
            elif c_idx == 2:
                cell.font = font_data
                cell.number_format = '0.00'
                cell.alignment = align_right
            elif c_idx == 4:
                cell.alignment = align_center
                if r_idx == 6:
                    cell.font = font_check_ok
                    cell.fill = fill_check_ok
                else:
                    cell.font = font_data

    # 4. 產出 異常情況工作表 Sheet
    ws_anom = wb.create_sheet(title='異常情況工作表')
    anom_headers = ['異常原因', '員工姓名', '身份證', '異常日期']
    ws_anom.append(anom_headers)
    for col_num in range(1, len(anom_headers) + 1):
        cell = ws_anom.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        
    for a_idx, anom in enumerate(anomalies, start=2):
        ws_anom.append([anom['異常原因'], anom['員工姓名'], anom['身份證'], anom['異常日期']])
        for c_idx in range(1, 5):
            cell = ws_anom.cell(row=a_idx, column=c_idx)
            cell.font = font_data
            cell.border = thin_border
            cell.alignment = align_center if c_idx >= 2 else align_left

    # 5. 所有工作表自動欄寬調整 (Auto Column Widths)
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                cell_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
                if cell_len > max_len:
                    max_len = cell_len
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # 存檔
    wb.save(OUTPUT_PAYROLL)
    print(f"成功產出薪資單工時統計表: {OUTPUT_PAYROLL}")


def generate_individual_payrolls(emp_dict, daily_records):
    """複製好寶寶薪資格式.xlsx 模版，依員工填入工時與基本資料並儲存為個人薪資檔案"""
    if not INPUT_TEMPLATE.exists():
        print(f"\n【警告】找不到個人薪資單模版檔案：{INPUT_TEMPLATE}")
        return

    print("\n=== 開始產出個人薪資單 Excel 檔案 ===")
    for nanny_name, emp_info in emp_dict.items():
        wb = openpyxl.load_workbook(INPUT_TEMPLATE)
        ws = wb.active
        
        emp_records = daily_records.get(nanny_name, {})
        
        # 累計當月各類別工時 (單位: 分鐘)
        norm_min = sum(r.get('正常工時', 0) for r in emp_records.values())
        ot_9_10_min = sum(r.get('平日加班_9_10h', 0) for r in emp_records.values())
        ot_11_12_min = sum(r.get('平日加班_11_12h', 0) for r in emp_records.values())
        weekday_transit_min = sum(r.get('平日轉場加時', 0) for r in emp_records.values())
        hol_0_2_min = sum(r.get('假日工時_0_2h', 0) for r in emp_records.values())
        hol_3_8_min = sum(r.get('假日工時_3_8h', 0) for r in emp_records.values())
        hol_9_12_min = sum(r.get('假日工時_9_12h', 0) for r in emp_records.values())
        weekend_transit_min = sum(r.get('假日轉場加時', 0) for r in emp_records.values())
        
        # 寫入個人基本資料 (紅底標題旁邊的黃色欄位)
        ws['C5'] = emp_info['姓名']
        ws['G5'] = emp_info['職稱']
        ws['E6'] = emp_info['身分證字號']
        ws['J6'] = emp_info['出生日期']
        
        # 寫入工時時數 (轉換為小時)
        ws['E8'] = norm_min / 60.0
        ws['E9'] = weekday_transit_min / 60.0
        ws['E16'] = ot_9_10_min / 60.0
        ws['E17'] = ot_11_12_min / 60.0
        ws['E18'] = hol_0_2_min / 60.0
        ws['E19'] = hol_3_8_min / 60.0
        ws['E20'] = hol_9_12_min / 60.0
        ws['E21'] = weekend_transit_min / 60.0
        
        # 檔名格式: <員工姓名>_<身份證字號>.xlsx
        out_filename = f"{emp_info['姓名']}_{emp_info['身分證字號']}.xlsx"
        out_path = DIR_PAYROLL / out_filename
        wb.save(out_path)
        print(f"成功產出個人薪資單: {out_filename}")


def main():
    print("=== 開始執行保母薪資單工時統計表與個人薪資單生成 ===")
    if not check_input_files_exist():
        input("\n請按 Enter 鍵結束...")
        return
        
    try:
        df_emp, df_hol, df_srv = load_input_data()
        emp_dict, daily_records, anomalies, raw_total_service_hours, all_dates = clean_and_process_data(df_emp, df_hol, df_srv)
        build_payroll_excel(emp_dict, daily_records, anomalies, raw_total_service_hours, all_dates)
        generate_individual_payrolls(emp_dict, daily_records)
        print(f"\n=== 全部分析與個人薪資單產出完成！異常紀錄筆數: {len(anomalies)} ===")
    except Exception as e:
        print(f"\n【執行發生錯誤】{e}")
    finally:
        input("\n請按 Enter 鍵結束...")

if __name__ == '__main__':
    main()
