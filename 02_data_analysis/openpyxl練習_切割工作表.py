import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill

target_dir = Path('02_data_analysis/outputs')
target_file_path = target_dir / 'sales.csv'


sales_df = pd.read_csv(target_file_path)

# 統計單位營業額（最新最優雅的寫法）
sales_by_dev = (
    sales_df
    .groupby(["業務單位"], as_index=False, dropna=False)
    .agg(sales_sum=("銷售金額", "sum"))
    .sort_values(by="sales_sum",ascending=False)
)

print("分群處理：")
print(sales_by_dev.loc[:])

# 切割部門資料
# 一行程式碼依「部門」切割成多個 DataFrame 物件（字典形式）
department_dfs = {dept: group.copy() for dept, group in sales_df.groupby("業務單位")}

# ----------------------------------------------------
# 🔍 如何使用切割後的 DataFrame？
# ----------------------------------------------------

# 查看切割出了哪些部門
print("切割出的部門：", list(department_dfs.keys()))

# 取得特定部門的 DataFrame（例如：財務部）
hr_df = department_dfs["業務3"]
print(hr_df.head())


# 依年+月份切割，例如2023-1月
def split_sales_by_period(df: pd.DataFrame) -> dict[pd.Period, pd.DataFrame]:
    """
    讀取 CSV 檔案，使用 Pandas 的 Period (時期) 物件以「月份」進行精準分組。
    回傳的字典 Key 為 Period 物件。
    """
    
    # 1. 將字串轉換為 datetime 時間型別
    df["銷售日期"] = pd.to_datetime(df["銷售日期"])
    
    # 2. 【核心優雅語法】直接轉換為「月 (Month)」等級的 Period 物件
    # 'M' 代表 Month，轉換後資料會變成如 2023-01, 2023-02 的 Period 型別
    df["銷售月份"] = df["銷售日期"].dt.to_period("M")
    
    # 3. 進行分組 (這裡的 p 是一個 Period 物件，而不是普通字串)
    period_dfs = {p: group.copy() for p, group in df.groupby("銷售月份")}
    # 加上 # type: ignore，強制靜音此行的型別警告，因為第三方套件的型別問題包袱太重
    return period_dfs # type: ignore

    
result_dict = split_sales_by_period(sales_df)

# 印出結果，看看 Period 物件的威力
for period_obj, df_group in result_dict.items():
    # period_obj 的型別是 <class 'pandas._libs.tslibs.period.Period'>
    print(f"📁 報表月份: {period_obj} | 筆數: {len(df_group)}")



# 將已建立的dataframe物件存入excel檔案的不同工作表

# 先過濾要儲存的欄位
# players_multi_conditions = players_multi_conditions.loc[:, ['firstname', 'lastname', 'ppg', 'apg', 'rpg']]

output_path = target_dir / '營業額分析.xlsx'


# 使用 pd.ExcelWriter 上下文管理器 (自動處理開啟與關閉)
with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    sales_df.to_excel(writer, sheet_name="原始資料", index=False)
    sales_by_dev.to_excel(writer, sheet_name="營業額by部門", index=False)

    # 走訪每個部門進行個別處理
    for dept_name, dept_df in department_dfs.items():
        dept_df.to_excel(writer, sheet_name=str(dept_name), index=False)
        print(f"【{dept_name}】資料筆數：{len(dept_df)}")

    # 走訪每個月份進行個別處理
    for period_obj, df_group in result_dict.items():
        # period_obj 的型別是 <class 'pandas._libs.tslibs.period.Period'>
        sheet_name = f'{period_obj.year}-{period_obj.month}'
        df_group.to_excel(writer, sheet_name=sheet_name, index=False)
        print(f"【{sheet_name}】資料筆數：{len(df_group)}")

    print("多工作表 Excel 建立完成！")

    # -------------------------------------------------------------
    # 設定 "sales over 2000萬" 的分組斑馬紋
    # -------------------------------------------------------------
    ws = writer.sheets["營業額by部門"]

    # 定義兩種背景顏色 (16進位 Hex 色碼，不需要加 #)
    fill_color_a = PatternFill(
        start_color="FF0000", end_color="FF0000", fill_type="solid"
    )  # 紅色 (純紅)

    fill_color_b = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )  # 灰色

    sales_target: int = 20000000

    # 從第 2 列開始走訪 (第 1 列是 Header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        # row[1] 對應第欄 (業務單位)

        # 當 fullname 改變時，切換填滿顏色
        if row[1].value >= sales_target:
            current_fill = fill_color_a
        else:
            current_fill = fill_color_b

        # 套用顏色至該列的所有儲存格
        for cell in row:
            cell.fill = current_fill


    print("多工作表 Excel 建立與動態斑馬紋格式設定完成！")